import openpyxl
# Crear un nuevo libro de Excel
workbook = openpyxl.Workbook()

# Seleccionar la hoja 
sheet = workbook.active
sheet.title = 'Listado'

# Agregar encabezados
sheet['A1'] = "Código"
sheet['B1'] = "Marca"
sheet['C1'] = "Modelo"
sheet['D1'] = "Precio"
sheet['E1'] = "Kilometraje"



# Lista de vehiculos 
Vehiculos = [
    ("001", "BMW", "Cx0012023", 4500000.00, 00),
    ("002", "Audi", "AU122023", 3000000.49,00),
    ("003", "Wolskvagen", "Wolk2024", 6500.79, 00),
    ("004", "Ferrari", "Xrt2022", 40000000.99,00),
    ("005", "Toyota", "Tacoma2023", 375000.29,00),

    
    # Agregar más Vehiculos aquí
]

# Agregar datos a la hoja
for row, data in enumerate(Vehiculos, start=2):
    sheet.cell(row=row, column=1, value=data[0])
    sheet.cell(row=row, column=2, value=data[1])
    sheet.cell(row=row, column=3, value=data[2])

 # Agregar más columnas dinámicamente
    for row, data  in enumerate(Vehiculos[3:], start=4):
        sheet.cell(row=row, column=4, value=data[4])

    for row, data  in enumerate(Vehiculos[4:], start=4):
        sheet.cell(row=row, column=5, value=data[4])



    #for col_index, value in enumerate(data[3:], start=4):
        #sheet.cell(row=row, column=col_index, value=value)

# Guardar el archivo Excel
workbook.save("vehículos.xlsx")

# Cerrar el archivo Excel
workbook.close()                                          